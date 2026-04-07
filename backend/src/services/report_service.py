"""
Report Generation Service - PDF & Excel
For on-premise deployment with no external dependencies
"""

import io
import csv
from datetime import datetime, timedelta
from typing import List, Dict, Optional
from dataclasses import dataclass

from sqlalchemy.orm import Session
from sqlalchemy import and_, or_, func

from ..models.lead import Lead
from ..models.user import User
from ..models import LeadAuditLog


@dataclass
class ReportConfig:
    """Report configuration"""

    title: str
    subtitle: str
    generated_by: str
    generated_at: datetime


class PDFReportService:
    """
    PDF Report Generation using ReportLab
    For Branch Manager reports
    """

    def __init__(self):
        self.page_width = 595  # A4 width in points
        self.page_height = 842  # A4 height in points
        self.margin = 50

    def _create_canvas(self, buffer: io.BytesIO):
        """Create PDF canvas"""
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4

        c = canvas.Canvas(buffer, pagesize=A4)
        return c

    def generate_lead_report(
        self,
        db: Session,
        start_date: datetime,
        end_date: datetime,
        branch_id: Optional[int] = None,
        generated_by: str = "System",
    ) -> bytes:
        """Generate PDF report for leads"""

        # Query leads
        query = db.query(Lead).filter(
            and_(Lead.created_at >= start_date, Lead.created_at <= end_date)
        )

        if branch_id:
            query = query.filter(Lead.branch_id == branch_id)

        leads = query.all()

        # Create PDF
        buffer = io.BytesIO()
        c = self._create_canvas(buffer)

        # Title page
        c.setFont("Helvetica-Bold", 24)
        c.drawString(self.margin, self.page_height - 100, "STBank Laos")

        c.setFont("Helvetica", 16)
        c.drawString(self.margin, self.page_height - 140, "Lead Generation Report")

        c.setFont("Helvetica", 12)
        c.drawString(
            self.margin,
            self.page_height - 170,
            f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
        )
        c.drawString(
            self.margin,
            self.page_height - 190,
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        )
        c.drawString(self.margin, self.page_height - 210, f"By: {generated_by}")

        # Summary
        c.setFont("Helvetica-Bold", 14)
        c.drawString(self.margin, self.page_height - 260, "Summary")

        c.setFont("Helvetica", 12)
        y = self.page_height - 290

        total_leads = len(leads)
        new_count = sum(1 for l in leads if l.status == "new")
        contacted = sum(1 for l in leads if l.status == "contacted")
        qualified = sum(1 for l in leads if l.status == "qualified")
        converted = sum(1 for l in leads if l.status == "converted")
        lost = sum(1 for l in leads if l.status == "lost")

        c.drawString(self.margin, y, f"Total Leads: {total_leads}")
        y -= 20
        c.drawString(self.margin, y, f"New: {new_count}")
        y -= 20
        c.drawString(self.margin, y, f"Contacted: {contacted}")
        y -= 20
        c.drawString(self.margin, y, f"Qualified: {qualified}")
        y -= 20
        c.drawString(self.margin, y, f"Converted: {converted}")
        y -= 20
        c.drawString(self.margin, y, f"Lost: {lost}")

        if total_leads > 0:
            y -= 30
            conversion_rate = (converted / total_leads) * 100
            c.drawString(self.margin, y, f"Conversion Rate: {conversion_rate:.1f}%")

        # Lead details on new page
        c.showPage()
        c.setFont("Helvetica-Bold", 14)
        c.drawString(self.margin, self.page_height - 50, "Lead Details")

        # Table header
        y = self.page_height - 90
        c.setFont("Helvetica-Bold", 8)
        headers = ["ID", "Name", "Phone", "Product", "Amount", "Status", "Date"]
        col_widths = [30, 80, 70, 60, 60, 50, 60]

        x = self.margin
        for i, header in enumerate(headers):
            c.drawString(x, y, header)
            x += col_widths[i]

        # Table rows
        c.setFont("Helvetica", 7)
        y -= 15

        for lead in leads[:50]:  # Max 50 per page
            if y < self.margin:
                c.showPage()
                y = self.page_height - 50

            x = self.margin
            row = [
                str(lead.id),
                (lead.full_name or "")[:20],
                (lead.phone or "")[:15],
                (lead.product or "")[:15],
                str(lead.amount or "-")[:10],
                (lead.status or "")[:10],
                (lead.created_at.strftime("%Y-%m-%d") if lead.created_at else "-")[:10],
            ]

            for i, cell in enumerate(row):
                c.drawString(x, y, cell[:20])
                x += col_widths[i]

            y -= 12

        c.save()
        buffer.seek(0)
        return buffer.getvalue()

    def generate_performance_report(
        self,
        db: Session,
        start_date: datetime,
        end_date: datetime,
        branch_id: Optional[int] = None,
        generated_by: str = "System",
    ) -> bytes:
        """Generate PDF performance report for branch managers"""

        # Query users (sales reps)
        users = db.query(User).filter(User.role == "sales_rep").all()

        # Calculate metrics per rep
        buffer = io.BytesIO()
        c = self._create_canvas(buffer)

        # Title
        c.setFont("Helvetica-Bold", 20)
        c.drawString(self.margin, self.page_height - 80, "STBank Laos")

        c.setFont("Helvetica", 14)
        c.drawString(self.margin, self.page_height - 110, "Sales Performance Report")

        c.setFont("Helvetica", 10)
        c.drawString(
            self.margin,
            self.page_height - 140,
            f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
        )

        # Metrics per user
        c.setFont("Helvetica-Bold", 12)
        c.drawString(
            self.margin, self.page_height - 180, "Performance by Sales Representative"
        )

        y = self.page_height - 210

        for user in users:
            # Get leads for this user
            leads = (
                db.query(Lead)
                .filter(
                    and_(
                        Lead.assigned_to == user.id,
                        Lead.created_at >= start_date,
                        Lead.created_at <= end_date,
                    )
                )
                .all()
            )

            if not leads:
                continue

            converted = sum(1 for l in leads if l.status == "converted")
            total = len(leads)
            rate = (converted / total * 100) if total > 0 else 0

            c.setFont("Helvetica-Bold", 10)
            c.drawString(self.margin, y, f"{user.email}")
            y -= 15

            c.setFont("Helvetica", 9)
            c.drawString(
                self.margin + 20,
                y,
                f"Total Leads: {total} | Converted: {converted} | Rate: {rate:.1f}%",
            )
            y -= 20

            if y < self.margin:
                c.showPage()
                y = self.page_height - 50

        c.save()
        buffer.seek(0)
        return buffer.getvalue()


class ExcelReportService:
    """
    Excel Report Generation using openpyxl
    For data export and analysis
    """

    def generate_leads_excel(
        self,
        db: Session,
        start_date: datetime,
        end_date: datetime,
        branch_id: Optional[int] = None,
    ) -> bytes:
        """Generate Excel report for leads"""

        # Query leads
        query = db.query(Lead).filter(
            and_(Lead.created_at >= start_date, Lead.created_at <= end_date)
        )

        if branch_id:
            query = query.filter(Lead.branch_id == branch_id)

        leads = query.all()

        # Create Excel
        output = io.BytesIO()

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Leads"

            # Headers
            headers = [
                "ID",
                "Full Name",
                "Phone",
                "Lao ID",
                "Product",
                "Amount (LAK)",
                "Preferred Time",
                "Status",
                "Assigned To",
                "Created At",
                "Converted At",
            ]

            header_font = Font(bold=True)
            header_fill = PatternFill(
                start_color="0066CC", end_color="0066CC", fill_type="solid"
            )

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Data rows
            for row_idx, lead in enumerate(leads, 2):
                assigned = db.query(User).filter(User.id == lead.assigned_to).first()

                ws.cell(row=row_idx, column=1, value=lead.id)
                ws.cell(row=row_idx, column=2, value=lead.full_name)
                ws.cell(row=row_idx, column=3, value=lead.phone)
                ws.cell(row=row_idx, column=4, value=lead.lao_id)
                ws.cell(row=row_idx, column=5, value=lead.product)
                ws.cell(
                    row=row_idx,
                    column=6,
                    value=float(lead.amount) if lead.amount else 0,
                )
                ws.cell(row=row_idx, column=7, value=lead.preferred_time)
                ws.cell(row=row_idx, column=8, value=lead.status)
                ws.cell(row=row_idx, column=9, value=assigned.email if assigned else "")
                ws.cell(
                    row=row_idx,
                    column=10,
                    value=(
                        lead.created_at.strftime("%Y-%m-%d %H:%M")
                        if lead.created_at
                        else ""
                    ),
                )
                ws.cell(
                    row=row_idx,
                    column=11,
                    value=(
                        lead.converted_at.strftime("%Y-%m-%d %H:%M")
                        if lead.converted_at
                        else ""
                    ),
                )

            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15

            wb.save(output)

        except ImportError:
            # Fallback to CSV if openpyxl not available
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)

            for lead in leads:
                writer.writerow(
                    [
                        lead.id,
                        lead.full_name,
                        lead.phone,
                        lead.lao_id,
                        lead.product,
                        lead.amount or 0,
                        lead.preferred_time,
                        lead.status,
                        (
                            lead.created_at.strftime("%Y-%m-%d %H:%M")
                            if lead.created_at
                            else ""
                        ),
                        (
                            lead.converted_at.strftime("%Y-%m-%d %H:%M")
                            if lead.converted_at
                            else ""
                        ),
                    ]
                )

            output = io.BytesIO(output.getvalue().encode())

        output.seek(0)
        return output.getvalue()

    def generate_audit_log_excel(
        self, db: Session, start_date: datetime, end_date: datetime
    ) -> bytes:
        """Generate Excel report for audit logs"""

        logs = (
            db.query(LeadAuditLog)
            .filter(
                and_(
                    LeadAuditLog.created_at >= start_date,
                    LeadAuditLog.created_at <= end_date,
                )
            )
            .all()
        )

        output = io.BytesIO()

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Audit Logs"

            headers = [
                "ID",
                "Lead ID",
                "User ID",
                "Action",
                "Old Status",
                "New Status",
                "IP Address",
                "Timestamp",
            ]

            header_font = Font(bold=True)
            header_fill = PatternFill(
                start_color="FF9900", end_color="FF9900", fill_type="solid"
            )

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill

            for row_idx, log in enumerate(logs, 2):
                ws.cell(row=row_idx, column=1, value=log.id)
                ws.cell(row=row_idx, column=2, value=log.lead_id)
                ws.cell(row=row_idx, column=3, value=log.user_id)
                ws.cell(row=row_idx, column=4, value=log.action)
                ws.cell(row=row_idx, column=5, value=log.old_status)
                ws.cell(row=row_idx, column=6, value=log.new_status)
                ws.cell(row=row_idx, column=7, value=log.ip_address)
                ws.cell(
                    row=row_idx,
                    column=8,
                    value=(
                        log.created_at.strftime("%Y-%m-%d %H:%M")
                        if log.created_at
                        else ""
                    ),
                )

            wb.save(output)

        except ImportError:
            # CSV fallback
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(
                [
                    "ID",
                    "Lead ID",
                    "User ID",
                    "Action",
                    "Old Status",
                    "New Status",
                    "IP Address",
                    "Timestamp",
                ]
            )

            for log in logs:
                writer.writerow(
                    [
                        log.id,
                        log.lead_id,
                        log.user_id,
                        log.action,
                        log.old_status,
                        log.new_status,
                        log.ip_address,
                        (
                            log.created_at.strftime("%Y-%m-%d %H:%M")
                            if log.created_at
                            else ""
                        ),
                    ]
                )

            output = io.BytesIO(output.getvalue().encode())

        output.seek(0)
        return output.getvalue()


class ReportService:
    """Combined report service"""

    def __init__(self):
        self.pdf_service = PDFReportService()
        self.excel_service = ExcelReportService()

    async def generate_branch_report(
        self,
        db: Session,
        start_date: datetime,
        end_date: datetime,
        branch_id: Optional[int],
        report_type: str,
        generated_by: str,
    ) -> bytes:
        """Generate report based on type"""

        if report_type == "pdf":
            return self.pdf_service.generate_lead_report(
                db, start_date, end_date, branch_id, generated_by
            )
        elif report_type == "excel":
            return self.excel_service.generate_leads_excel(
                db, start_date, end_date, branch_id
            )
        else:
            raise ValueError(f"Unknown report type: {report_type}")

    async def generate_audit_report(
        self, db: Session, start_date: datetime, end_date: datetime, report_type: str
    ) -> bytes:
        """Generate audit report"""

        if report_type == "excel":
            return self.excel_service.generate_audit_log_excel(db, start_date, end_date)
        else:
            raise ValueError(f"Unknown report type: {report_type}")
